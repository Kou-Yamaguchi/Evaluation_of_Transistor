"""
オリジナル
2022/04/19
H. Uryu, 1521512@ed.tus.ac.jp(2023卒)

編集者
R. Kaneko 1519032@ed.tus.ac.jp(2025卒)
K. Tomiyoshi 1522529@ed.tus.ac.jp(2024卒)
M. Taniguchi 1521536@ed.tus.ac.jp(2023卒)

2023/10/26
K. Yamaguchi 1520107@ed.tus.ac.jp(2026卒)
トランジスタ測定用（Vd固定-Vgスイープ）に改良

詳しくはGoogle Drive内の"更新情報"を参照
"""
#default設定
d_interval = 0.1#[s]
d_V_min = -0.8#[V]
d_V_max = 0.8#[s]
d_V_step = 0.1#[V]
d_loop = 1#回
d_folderpath = 'C:/Users/higuchi/Desktop/IVスイープ'
d_x_label = "Voltage [V]"
d_y_label = "Current [A]"

import matplotlib.pyplot as plt
import os
import pyvisa as visa
import re
import threading
import time
import tkinter as tk
import numpy as np 
from decimal import Decimal
from tkinter import filedialog
from tkinter import ttk
from numpy import format_float_scientific as sci

rm = visa.ResourceManager(r'C:\WINDOWS\system32\visa64.dll')
#dev_=rm.list_resources()
devs = []
dev0 = rm.open_resource('GPIB0::1::INSTR')#ゲート用
dev1 = rm.open_resource('GPIB1::1::INSTR')#ドレイン用
devs.append(dev0)
devs.append(dev1)
for dev in devs:
    dev.timeout = 5000
    print(dev.query('*IDN?'))

#送信コマンド
def write(dev, command):
    dev.write(command)

#受信コマンド
def query(dev, command):
    dev.query(command)

#フォルダ選択
def set_folder_func():
    dir = 'C:\\'
    folder_path = filedialog.askdirectory(initialdir = dir)
    textbox["folderpath"].delete(0, tk.END)
    textbox["folderpath"].insert(tk.END, folder_path)  

#グラフ
def graph(x_list, y_list, plot, scatter, x_label, y_label):
    def para(dic):
        return {f'{k1}.{k2}' : v for k1, d in dic.items() for k2, v in d.items()} 
    config = {
        "font" :{
            "family":"Times New Roman",
            "size":14
            },
        "xtick" :{
                "direction":"in",
                "top":True,
                "major.width":1.2,
                "labelsize":20.0
            },
        "ytick" :{
                "direction":"in",
                "right":True,
                "major.width":1.2,
                "labelsize":20.0
            },
        "axes" :{
            "linewidth":1.2,
            "labelpad":10
            },
        
        "figure" :{
            "dpi":150
                }
        }
    
    plt.rcParams.update(para(config))
    
    fig=plt.figure()
    ax=fig.add_subplot()

    if plot == True:
        ax.plot(x_list, y_list)
    if scatter == True:
        ax.scatter(x_list, y_list)
        
    ax.set_xlabel(x_label)
    ax.set_ylabel(y_label)  
    plt.show()

#強制停止
def stop_func():
    global stop_flag
    stop_flag = True
    swrite("測定中断")

#IVスイープ
def measure(start, stop, step, constV, sweep_times, interval, plot, scatter, value, sweepVoltage): 
    global stop_flag
    
    for i in range(len(devs)):
        if i == sweepVoltage:
            write(devs[i], "SN"+str(start)+","+str(stop)+","+str(step))#SNstart,stop,step(リニアスイープ測定設定)
        else:
            write(devs[i], "SOV"+str(constV))#SNconstV(DC電圧発生)
        write(devs[i], "OPR")#出力

    print("測定実行")
    for _ in range(sweep_times):
        if stop_flag == True:
            for dev in devs:
                write(dev, "SWSP")#実行中のスイープを停止
            break
        
        for dev in devs:
            write(dev, "*TRG")#測定実行
        time.sleep(interval)

        A_G=devs[0].query("N?")
        A_G_ = float(A_G[3:-2])
        A_Glist.append(A_G_)
        
        V_G=devs[0].query("SOV?")  
        V_G_ = float(V_G[3:-2])
        V_Glist.append(V_G_)

        A_D=devs[1].query("N?")
        A_D_ = float(A_D[3:-2])
        A_Dlist.append(A_D_)
        
        V_D=devs[1].query("SOV?")  
        V_D_ = float(V_D[3:-2])
        V_Dlist.append(V_D_)


        if value == True:
            if not A_G_ == 0:
                print(f"{V_G_:.6f} V\r\n{sci(A_G_, precision = 6, exp_digits = 2)} A\r\n{sci(V_G_/A_G_, precision = 6, exp_digits = 2)} Ω\r\n")
            else: 
                print(f"{V_G_:.6f} V\r\n{sci(A_G_, precision = 6, exp_digits = 2)} A\r\nError Ω\r\n")
        if plot == True or scatter == True:
            graph(V_Glist, A_Glist, plot, scatter, "Gate Voltage [Vg]", "Gate Current [Ig]")
            graph(V_Dlist, A_Dlist, plot, scatter, "Drain Voltage [Vd]", "Drain Current [Id]")
        
    for dev in devs:
            write(dev, "SBY")
    
#ファイル出力
def output(filepath, x1_list, y1_list, x2_list, y2_list, extension_index):
    def output_txt():
        with open(filepath, 'w') as data:
            for w, x, y, z in zip(x1_list, y1_list, x2_list, y2_list):
                data.write(f"{str(w)} {str(x)} {str(y)} {str(z)}\n")    
    
    def output_csv():
        import csv
        with open(filepath, 'w', newline="") as data:
            writer = csv.writer(data)
            for w, x, y, z in zip(x1_list, y1_list, x2_list, y2_list):
                writer.writerow([w, x, y, z])
                
    def output_xlsx():
        from openpyxl import Workbook
        from openpyxl import load_workbook
        
        wb = Workbook()
        wb.save(filepath)
        wb = load_workbook(filepath)
        ws = wb['Sheet']
        ws = wb.active
        
        ws.cell(1, 1, "Gate Voltage")
        ws.cell(1, 2, "Gate Current")
        ws.cell(1, 3, "Drain Voltage")
        ws.cell(1, 4, "Drain Current")
        
        for i, (x1_val, y1_val, x2_val, y2_val) in enumerate(zip(x1_list, y1_list, x2_list, y2_list), 1):
            ws.cell(i+1, 1, x1_val)#gate Voltage
            ws.cell(i+1, 2, y1_val)#gate Current
            ws.cell(i+1, 3, x2_val)#Drain Current
            ws.cell(i+1, 4, y2_val)#Drain Current
            
        wb.save(filepath)
        wb.close()    
        
    if extension_index == 0:
        output_txt()          
    if extension_index == 1:
        output_csv()
    if extension_index == 2:
        output_xlsx()

#抵抗値表示
def resistance(x_list, y_list, x_label, y_label):
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    
    x = np.array([x_list, np.ones(len(x_list))])
    x = x.T
    a, b = np.linalg.lstsq(x, y_list)[0]
    
    def para(dic):
        return {f'{k1}.{k2}' : v for k1, d in dic.items() for k2, v in d.items()} 
    config = {
        "font" :{
            "family":"Times New Roman",
            "size":14
            },
        "xtick" :{
                "direction":"in",
                "top":True,
                "major.width":1.2,
                "labelsize":20.0
            },
        "ytick" :{
                "direction":"in",
                "right":True,
                "major.width":1.2,
                "labelsize":20.0
            },
        "axes" :{
            "linewidth":1.2,
            "labelpad":10
            },
        
        "figure" :{
            "figsize":(16, 9),
            "dpi":60
                }
        }
    plt.rcParams.update(para(config))    

    fig = plt.figure()
    ax = fig.add_subplot()
    ax.plot(x_list, y_list, "ro")
    ax.plot(x, (a*x+b), "b")
    ax.set_xlabel(x_label)
    ax.set_ylabel(y_label)

    root1 = tk.Toplevel()
    root1.title("Resistance")
    root1.geometry("850x540")
    root1.resizable(False, False)
    root1.lift()
    
    canvas = FigureCanvasTkAgg(fig, master = root1)
    canvas.get_tk_widget().pack()
    
    def create_label(config):
        for var in config:
            label[var[0]] = tk.Label(root1, text = var[0], background = '#B0E0E6', font = ('Times New Roman', "20"))
            label[var[0]].place(x = var[1] + var[2]*var[5], y = var[3] + var[4]*var[5])
    label = {}
    label_list = [[f'Resistance: {sci(1/a, precision = 4, exp_digits = 4)} [Ω]'], [f'y = {sci(1/a, precision = 4, exp_digits = 4)}x + {sci(b, precision = 4, exp_digits = 4)}'],]
    label_params = [[25, 0, 10, 30],[400, 0, 10, 30]]
    label_config = [[tag_] + con + [i] for tag, con in zip(label_list, label_params) for i, tag_ in enumerate(tag)]
    create_label(label_config)

    print(f'Resistance: {sci(1/a, precision = 4, exp_digits = 4)} [Ω]', f'y = {sci(a, precision = 4, exp_digits = 4)}x + {sci(b, precision = 4, exp_digits = 4)}')
    
#実行関数
def run_func():
    global V_Glist, A_Glist, V_Dlist, A_Dlist, stop_flag
    V_Glist, A_Glist, V_Dlist, A_Dlist =[], [], [], []
    stop_flag = False
    interval = float(spinbox["interval"].get())#[s]
    loop = float(spinbox["loop"].get())#ループ回数
    V_min = Decimal(spinbox["V_min"].get())#[V]
    V_max = Decimal(spinbox["V_max"].get())#[V]
    V_step = Decimal(spinbox["V_step"].get())#[V]
    V_const = Decimal(spinbox["V_const"].get())#[V]
    V_start = 0#[V]

    chk0 = checkbutton['ファイルに出力する'].get()
    chk1 = checkbutton['プロットを表示する'].get()
    chk2 = checkbutton['散布図を表示する'].get()
    chk3 = checkbutton['抵抗値を表示する'].get()
    chk4 = checkbutton['測定値を表示する'].get()

    extension_box_index = combobox["ext"].current()
    extension = combobox["ext"].get()
    sweepmode = combobox["sweepmode"].current()
    sweepVoltage = combobox["sweepVoltage"].current()
    
    step_chk = abs(V_max-V_min)/V_step
    
    for i in range(len(devs)):
        write(devs[i], "*RST")#初期化
        write(devs[i], "M1")#トリガーモード HOLD
        write(devs[i], "OH1")#ヘッダON
        write(devs[i], "VF")#電圧発生
        write(devs[i], "F2")#電流測定
        if i == sweepVoltage:
            write(devs[i], "MD2")#DCスイープモード
        else:
            write(devs[i], "MD0")#DCモード
        write(devs[i], "R0")#オートレンジ

    
    #エラーチェック
    if chk0 == True:
        folderpath = textbox["folderpath"].get()
        filename = textbox["filename"].get()
        
        if not os.path.exists(folderpath):
            swrite("無効なフォルダーパスです")
            return
        if filename == "" and extension == ".csv":
            swrite("ファイル名を入力して下さい")
            return
        filename = re.sub(r'[\\/:*?"<>|]+', '_', filename)
        filepath = folderpath +'/' + filename + extension
    
    if loop.is_integer() == False:
        swrite("ループ回数は整数値を設定して下さい")
        return
    else:
        loop = int(loop)
    
    if float(step_chk).is_integer() == False:
        swrite("ステップ数が整数になるように設定して下さい")
        return
    
    swrite("測定中")

    #測定実行
    sweeptimes1 = int(step_chk)+1
    sweeptimes2 = int((V_max-V_start)/V_step)+1
    sweeptimes3 = int((V_start-V_min)/V_step)+1    
    
    #双方向スイープ
    if sweepmode == 0:
        for i in range(loop): 
            swrite(f"測定中: ループ {i+1}/{loop}")
            measure(V_start, V_max, V_step, V_const, sweeptimes2, interval, chk1, chk2, chk4, sweepVoltage)        
            measure(V_max, V_min, V_step, V_const, sweeptimes1, interval, chk1, chk2, chk4, sweepVoltage)
            measure(V_min, V_start, V_step, V_const, sweeptimes3, interval, chk1, chk2, chk4, sweepVoltage)
        swrite("測定終了")
        
    #単方向スイープ
    if sweepmode == 1:
        for i in range(loop):
            swrite(f"測定中: ループ {i+1}/{loop}")
            measure(V_min, V_max, V_step, V_const, sweeptimes1, interval, chk1, chk2, chk4, sweepVoltage)
            measure(V_max, V_min, V_step, V_const, sweeptimes1, interval, chk1, chk2, chk4, sweepVoltage)
        swrite("測定終了")
    
    #単方向(折り返しなし)
    if sweepmode == 2:
        measure(V_min, V_max, V_step, V_const, sweeptimes1, interval, chk1, chk2, chk4, sweepVoltage)
        swrite("測定終了")

    #ファイルに出力する場合      
    if chk0 == True:
        output(filepath, V_Glist, A_Glist, V_Dlist, A_Dlist, extension_box_index)
        
    #抵抗値の表示
    if chk3 == True:
        resistance(V_Glist, A_Glist, "Gate Voltage Vg [V]", "Gate Current Ig [A]")
        resistance(V_Glist, A_Dlist, "Gate Voltage Vg [V]", "Drain Current Id [A]")

def exc_run_func():
    try:      
        t1 = threading.Thread(target = run_func)
        t1.start()

    except:
        swrite("予期せぬエラーです")

#ウィンドウ
root = tk.Tk()
root.title("I-V Sweep ver1.4")
root.geometry("430x325")#横×縦
root.resizable(False, False)#ウィンドウサイズをフリーズ
root.lift()#最前面に表示

#ラベル
def create_label(config):
            for var in config: 
                if var[5] == True:
                    label[var[0]] = tk.Label(text= var[0], background= '#B0E0E6')
                else:
                    label[var[0]] = tk.Label(text= var[0])
                label[var[0]].place(x=var[1] + var[2]*var[6], y= var[3] + var[4]*var[6])

label = {} 
label_list = [['保存先のフォルダ', 'ファイル名を入力'],
              ['V_min [V]', 'V_max [V]', 'V_step [V]','V_const [V]', '遅延 [s]', 'ループ回数'],
              ['※折り返し無しの場合、無効'],
              ['ファイル形式'],
              ['測定モード','スイープ電圧'],]
#x = a+bx, y=c+dxを満たす[a, b, c, d] + background   
label_params = [[25, 0, 10, 30, True],
                [40, 0, 75, 25, False],
                [40, 0, 220, 0, False],
                [290, 0, 40, 0, True],
                [205, 0, 185, 25, True],]
label_config = [[tag_] + con + [i] for tag, con in zip(label_list, label_params) for i, tag_ in enumerate(tag)]
create_label(label_config)

#テキストボックス
def create_textbox(config):
    for key, var in config.items():
        textbox[key] = ttk.Entry(width= var[0])
        textbox[key].place(x= var[1], y= var[2])
        textbox[key].insert(0, var[3])
        
textbox = {}
textbox_config = {
    #{tag :[wid, x, y, init]}
    "folderpath" :[38, 120, 10, d_folderpath],
    "filename" :[25, 120, 40, ""],
    }  
create_textbox(textbox_config)

#スピンボックス
def create_spinbox(config):
    for i, (key, var) in enumerate(config.items()):
        spinbox[key] = ttk.Spinbox(
            root, 
            width = 7,
            format = '%3.1f',
            from_ = var[0],
            to = var[1],
            increment = var[2],
            )            
        spinbox[key].place(x= 125, y= 75 + 25*i)
        spinbox[key].insert(0, var[3])

spinbox = {}
spinbox_config = {
    #{tag :[min, max, step, init]}
    "V_min" :[-30.0, 30.0, 0.1, d_V_min],
    "V_max" :[-30.0, 30.0, 0.1, d_V_max],
    "V_step" :[-30.0, 30.0, 0.1, d_V_step],
    "V_const":[-30.0, 30.0, 0.1, 1],
    "interval" :[0.0, 10000.0, 0.1, d_interval],
    "loop":[1, 10000, 1, d_loop],
    }
create_spinbox(spinbox_config)

#チェックボタン
def create_checkbutton(config):
    for i, (key, var) in enumerate(config.items()):
        checkbutton[key] = tk.BooleanVar()
        checkbutton[key].set(var)
        chk = ttk.Checkbutton(
            root,
            variable = checkbutton[key],
            text = key
            )
        chk.place(x= 230, y= 75 + 20*i)

checkbutton = {}
checkbutton_config = {
    #[text :bln]
    'ファイルに出力する' :True,
    'プロットを表示する' :False,
    '散布図を表示する' :False,
    '抵抗値を表示する' :True,
    '測定値を表示する':False,
    }
        
create_checkbutton(checkbutton_config)

#ボタン
def create_button(config):
    for key, var in config.items():
        button[key] = ttk.Button(
            root,
            text = key,
            width = var[0],
            padding = [var[1], var[2]],
            command = var[5],
            )
        button[key].place(x= var[3], y= var[4])
        
button = {}
button_config = {
    #{tag :[wid, pad_EW, pad_NS, x, y, command]}
    "参照": [8, 0, 0, 360, 9, set_folder_func],
    "実行": [12, 0, 10, 125, 250, exc_run_func],
    "強制終了": [12, 0, 10, 225, 250, stop_func],
    }
create_button(button_config)

#プルダウンリスト
def create_combobox(config):
    for key, var in config.items():
        combobox[key] = ttk.Combobox(
            root,
            width = var[0],
            justify = "left", 
            state = "readonly",
            values = var[1],
            )
        combobox[key].place(x= var[2], y= var[3])
        combobox[key].current(var[4])
        
combobox = {}
combobox_config = {
    #tag :[wid, [values], x, y, init]
    "ext": [4, [".txt", ".csv", ".xlsx"], 360, 40, 2],
    "sweepmode": [18, ["双方向スイープ", "単方向スイープ", "単方向(折り返し無し)"], 275, 185, 0],
    "sweepVoltage": [18, ["ゲート電圧(Vg)","ドレイン電圧(Vd)"], 275, 210, 0],     
    }
create_combobox(combobox_config)

statusbar = tk.Label(root, text = "", bd = 1, relief = tk.SUNKEN, anchor = tk.W)
statusbar.pack(side = tk.BOTTOM, fill = tk.X)
def swrite(text):
    statusbar["text"] = text

root.mainloop()